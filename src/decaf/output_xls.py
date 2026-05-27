"""Excel output for tax report — one sheet per quadro."""

from __future__ import annotations

from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from decaf.country_codes import iso_to_ade_country_code
from decaf.forex import THRESHOLD_EUR
from decaf.models import TaxReport

_HEADER_FONT = Font(bold=True, size=11)
_HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
_MONEY_FMT = "#,##0.00"
_THIN_BORDER = Border(
    bottom=Side(style="thin", color="B0B0B0"),
)


def _strip_acquired(description: str) -> str:
    """Drop trailing ' (acquired YYYY-MM-DD)' — date is in its own column."""
    idx = description.find(" (acquired ")
    return description[:idx] if idx >= 0 else description


def write_xls(report: TaxReport, path: Path) -> None:
    """Write the tax report as an Excel workbook."""
    wb = Workbook()

    summary_ws = wb.active
    assert isinstance(summary_ws, Worksheet)
    _write_summary(summary_ws, report)
    _write_precompilata(wb.create_sheet("Precompilata"), report)
    _write_rw(wb.create_sheet("Quadro RW"), report)
    _write_rt(wb.create_sheet("Quadro RT"), report)
    _write_rl(wb.create_sheet("Quadro RL"), report)
    _write_forex(wb.create_sheet("Analisi Soglia Valutaria"), report)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))


def _write_summary(ws: Worksheet, report: TaxReport) -> None:
    ws.title = "Riepilogo"

    ws.append(["Report Fiscale Italiano", "", f"Anno fiscale {report.tax_year}"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])

    ws.append(["Informazioni Conto"])
    ws["A3"].font = Font(bold=True, size=12)
    ws.append(["ID Conto", report.account.account_id])
    ws.append(["Intestatario", report.account.holder_name])
    ws.append(["Broker", report.account.broker_name])
    ws.append(["Paese", report.account.country])
    ws.append(["Valuta base", report.account.base_currency])
    ws.append(["Data apertura", report.account.date_opened.isoformat()])
    ws.append([])

    ws.append(["Riepilogo Fiscale"])
    ws["A11"].font = Font(bold=True, size=12)
    ws.append(["", "Importo (EUR)"])
    ws["B12"].font = _HEADER_FONT

    ws.append(["IVAFE totale (Quadro RW)", float(report.total_ivafe)])
    ws["B13"].number_format = _MONEY_FMT

    ws.append(["Plusvalenze nette (Quadro RT)", float(report.net_capital_gain_loss)])
    ws["B14"].number_format = _MONEY_FMT

    ws.append(["Redditi lordi (Quadro RL)", float(report.total_gross_interest_eur)])
    ws["B15"].number_format = _MONEY_FMT

    ws.append(["Ritenute estere (Quadro RL)", float(report.total_wht_eur)])
    ws["B16"].number_format = _MONEY_FMT

    ws.append([])
    ws.append(["Soglia Valutaria"])
    ws["A18"].font = Font(bold=True, size=12)
    ws.append(["Soglia (EUR)", float(THRESHOLD_EUR)])
    ws["B19"].number_format = _MONEY_FMT
    ws.append(["Superata", "SI" if report.forex_threshold_breached else "NO"])
    ws.append(["Giorni lavorativi consecutivi", report.forex_max_consecutive_days])
    if report.forex_first_breach_date:
        ws.append(["Data prima violazione", report.forex_first_breach_date.isoformat()])

    if report.rsu_vest_count:
        ws.append([])
        header_row = ws.max_row + 1
        ws.append(["Controllo coerenza RSU (art. 9 c. 4 TUIR)"])
        ws.cell(row=header_row, column=1).font = Font(bold=True, size=12)
        ws.append(["Vest events nell'anno", report.rsu_vest_count])
        ws.append(
            [
                "Reddito RSU tassato (EUR)",
                float(report.rsu_income_eur),
            ]
        )
        ws.cell(row=ws.max_row, column=2).number_format = _MONEY_FMT
        note_row = ws.max_row + 1
        ws.append(
            [
                "Cross-check: deve essere sottoinsieme del punto 1 CU "
                '"Redditi di lavoro dipendente". Differenza = stipendio + bonus.'
            ]
        )
        ws.cell(row=note_row, column=1).font = Font(italic=True, size=9)
        ws.merge_cells(start_row=note_row, end_row=note_row, start_column=1, end_column=4)

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 25


def _write_precompilata(ws: Worksheet, report: TaxReport) -> None:
    """One sheet with everything the user needs to fill into the AdE precompilata.

    Layout: for each quadro, a small block with rigo/colonna/valore/origine.
    """
    ws.title = "Precompilata"

    bold_blue = Font(bold=True, size=12, color="1F4E78")
    italic_dim = Font(italic=True, size=9, color="595959")

    ws.append([f"Per la dichiarazione precompilata (anno fiscale {report.tax_year})"])
    ws["A1"].font = Font(bold=True, size=14, color="1F4E78")
    ws.append(["Mappa decaf -> Modello Redditi PF. Valori in EUR salvo dove indicato."])
    ws["A2"].font = italic_dim
    ws.append([])

    def block_header(title: str, subtitle: str = "") -> None:
        r = ws.max_row + 1
        ws.append([title])
        ws.cell(row=r, column=1).font = bold_blue
        if subtitle:
            ws.append([subtitle])
            ws.cell(row=r + 1, column=1).font = italic_dim
        # column headers
        hr = ws.max_row + 1
        ws.append(["Rigo", "Colonna", "Valore EUR", "Origine"])
        for col in range(1, 5):
            c = ws.cell(row=hr, column=col)
            c.font = _HEADER_FONT
            c.fill = _HEADER_FILL
            c.alignment = Alignment(horizontal="center")

    # --- Quadro RW ---
    if report.rw_lines:
        # RW uses custom headers (per-row layout, not rigo/colonna)
        title_row = ws.max_row + 1
        ws.append(["Quadro RW - un modulo per ogni riga"])
        ws.cell(row=title_row, column=1).font = bold_blue
        ws.append(["Colonne fisse: col.1=1 (proprietà), col.5=100%, col.6=1 (valore di mercato)."])
        ws.cell(row=title_row + 1, column=1).font = italic_dim
        hdr_row = ws.max_row + 1
        ws.append(
            [
                "Modulo",
                "col.3 Cod. / col.4 Stato (AdE)",
                "col.7 Val. iniz. / col.8 Val. fin. / col.10 Giorni",
                "col.30 IVAFE / Note",
            ]
        )
        for col in range(1, 5):
            c = ws.cell(row=hdr_row, column=col)
            c.font = _HEADER_FONT
            c.fill = _HEADER_FILL
            c.alignment = Alignment(horizontal="center")
        for idx, line in enumerate(report.rw_lines, start=1):
            note = (
                f"{line.ivafe_due:,.2f}   |   "
                f"{line.symbol} {_strip_acquired(line.long_description)}"
            )
            ade_country = iso_to_ade_country_code(line.country) or "?"
            ws.append(
                [
                    f"RW{idx}",
                    f"{line.codice_investimento} / {ade_country} ({line.country})",
                    (
                        f"{line.initial_value_eur:,.2f} / "
                        f"{line.final_value_eur:,.2f} / {line.days_held}"
                    ),
                    note,
                ]
            )
        ws.append([])

    # --- Quadro RT ---
    if report.rt_lines:
        block_header(
            "Quadro RT - Sez. II-A (imposta sostitutiva 26%)",
            "Partecipazioni non qualificate. Un solo rigo aggregato (RT11). "
            "Minus pregresse: RT13 (manuale).",
        )
        ws.append(
            [
                "RT11",
                "col.1 - Totale corrispettivi",
                float(report.rt_total_proceeds_eur),
                f"somma proceeds_eur ({len(report.rt_lines)} righe)",
            ]
        )
        ws.cell(row=ws.max_row, column=3).number_format = _MONEY_FMT
        ws.append(
            [
                "RT11",
                "col.2 - Totale costi",
                float(report.rt_total_cost_basis_eur),
                f"somma cost_basis_eur ({len(report.rt_lines)} righe)",
            ]
        )
        ws.cell(row=ws.max_row, column=3).number_format = _MONEY_FMT
        net = report.net_capital_gain_loss
        ws.append(
            [
                "",
                "Differenza (plus/minus netta)",
                float(net),
                "calcolata in automatico dal software AdE",
            ]
        )
        ws.cell(row=ws.max_row, column=3).number_format = _MONEY_FMT
        if net > 0:
            ws.append(
                [
                    "",
                    "Imposta sostitutiva 26%",
                    float(net * Decimal("0.26")),
                    "calcolata in automatico dal software AdE",
                ]
            )
            ws.cell(row=ws.max_row, column=3).number_format = _MONEY_FMT
        ws.append([])

    # --- Quadro RL + Quadro RM (mutually exclusive) ---
    if report.rl_lines:
        from decaf.quadro_rl import aggregate_rl_for_rm31

        gross = report.total_gross_interest_eur
        wht = report.total_wht_eur
        rm_imposta = gross * Decimal("0.26")
        groups = aggregate_rl_for_rm31(report.rl_lines)

        # --- Opzione B (RM31) — one rigo per (stato, tipo) ---
        title_row = ws.max_row + 1
        ws.append(["Opzione B (consigliata) - Quadro RM rigo RM31"])
        ws.cell(row=title_row, column=1).font = bold_blue
        ws.append(
            [
                "Sez. II-A imposta sostitutiva 26%. Niente Quadro CE. "
                "Un rigo per coppia (stato estero, tipo reddito)."
            ]
        )
        ws.cell(row=title_row + 1, column=1).font = italic_dim
        hdr_row = ws.max_row + 1
        ws.append(
            [
                "Rigo",
                "col.1 Tipo",
                "col.2 Stato (AdE)",
                "col.3 Lordo EUR",
                "col.4 Aliq.",
                "col.8 Imposta EUR",
                "Origine",
            ]
        )
        for col in range(1, 8):
            c = ws.cell(row=hdr_row, column=col)
            c.font = _HEADER_FONT
            c.fill = _HEADER_FILL
            c.alignment = Alignment(horizontal="center")
        for idx, g in enumerate(groups, start=1):
            ade = iso_to_ade_country_code(g["stato"]) or "?"
            imposta = g["gross_eur"] * Decimal("0.26")
            ws.append(
                [
                    f"RM31 (#{idx})",
                    g["tipo"],
                    f"{ade} ({g['stato']})",
                    float(g["gross_eur"]),
                    26,
                    float(imposta),
                    f"{g['label']} ({g['count']} entries)",
                ]
            )
            ws.cell(row=ws.max_row, column=4).number_format = _MONEY_FMT
            ws.cell(row=ws.max_row, column=6).number_format = _MONEY_FMT
        ws.append(
            [
                "TOTALE",
                "",
                "",
                float(gross),
                "",
                float(rm_imposta),
                f"{len(groups)} righi - WHT estera EUR {wht:,.2f} NON recuperabile",
            ]
        )
        ws.cell(row=ws.max_row, column=4).number_format = _MONEY_FMT
        ws.cell(row=ws.max_row, column=4).font = _HEADER_FONT
        ws.cell(row=ws.max_row, column=6).number_format = _MONEY_FMT
        ws.cell(row=ws.max_row, column=6).font = _HEADER_FONT
        ws.append([])

        # --- Opzione A (RL2) — compact ---
        title_row = ws.max_row + 1
        ws.append(["Opzione A (alternativa) - Quadro RL rigo RL2 + Quadro CE"])
        ws.cell(row=title_row, column=1).font = bold_blue
        ws.append(
            [
                "IRPEF marginale + credito d'imposta estera. Zona grigia per "
                "dividendi non qualificati (Cass. 35454/2022)."
            ]
        )
        ws.cell(row=title_row + 1, column=1).font = italic_dim
        hdr_row = ws.max_row + 1
        ws.append(["Rigo", "Colonna", "Valore EUR", "Origine"])
        for col in range(1, 5):
            c = ws.cell(row=hdr_row, column=col)
            c.font = _HEADER_FONT
            c.fill = _HEADER_FILL
            c.alignment = Alignment(horizontal="center")
        ws.append(
            [
                "RL2",
                "col.1 - Tipo reddito (dropdown)",
                "manuale",
                "es. B (dividendi non qualif.)",
            ]
        )
        ws.append(
            [
                "RL2",
                "col.2 - Redditi (lordo)",
                float(gross),
                "somma gross_amount_eur",
            ]
        )
        ws.cell(row=ws.max_row, column=3).number_format = _MONEY_FMT
        ws.append(
            [
                "RL2",
                "col.3 - Ritenute",
                float(wht),
                "somma wht_amount_eur (credito art. 165 -> CE)",
            ]
        )
        ws.cell(row=ws.max_row, column=3).number_format = _MONEY_FMT
        ws.append([])

        # Comparison block
        cmp_hdr_row = ws.max_row + 1
        ws.append(["Confronto imposta italiana - scegli UNA via"])
        ws.cell(row=cmp_hdr_row, column=1).font = bold_blue
        ws.append(
            [
                "MUTUAMENTE ESCLUSIVE (circ. 165/E §6). Aliquote IRPEF nominali "
                "senza addizionali (~1-2.5%)."
            ]
        )
        ws.cell(row=cmp_hdr_row + 1, column=1).font = italic_dim
        hdr = ws.max_row + 1
        ws.append(
            [
                "Scenario",
                "Imposta italiana EUR",
                "Totale Italia+WHT EUR",
                "Calcolo",
                "Convenienza",
            ]
        )
        for col in range(1, 6):
            c = ws.cell(row=hdr, column=col)
            c.font = _HEADER_FONT
            c.fill = _HEADER_FILL
        ws.append(
            [
                "RM31 (sost. 26%)",
                float(rm_imposta),
                float(rm_imposta + wht),
                f"{gross:,.2f} x 26% (WHT {wht:,.2f} persa)",
                "",
            ]
        )
        ws.cell(row=ws.max_row, column=2).number_format = _MONEY_FMT
        ws.cell(row=ws.max_row, column=3).number_format = _MONEY_FMT
        for rate, label in [
            (Decimal("0.23"), "RL+CE @23% (fino 28k)"),
            (Decimal("0.35"), "RL+CE @35% (28-50k)"),
            (Decimal("0.43"), "RL+CE @43% (oltre 50k)"),
        ]:
            irpef = gross * rate
            credito = min(wht, irpef)
            netta = irpef - credito
            adv = "PIU' CONVENIENTE" if netta < rm_imposta else ""
            ws.append(
                [
                    label,
                    float(netta),
                    float(netta + wht),
                    f"{gross:,.2f} x {int(rate * 100)}% - credito {credito:,.2f}",
                    adv,
                ]
            )
            ws.cell(row=ws.max_row, column=2).number_format = _MONEY_FMT
            ws.cell(row=ws.max_row, column=3).number_format = _MONEY_FMT
        be = (Decimal("0.26") + wht / gross) if gross else Decimal(0)
        ws.append(
            [f"Break-even aliquota marginale: {be * 100:.1f}% (sotto -> RL+CE; sopra -> RM31)"]
        )
        ws.cell(row=ws.max_row, column=1).font = italic_dim
        ws.append([])

        # Recommendation
        rec_row = ws.max_row + 1
        ws.append(["Raccomandazione: RM31"])
        ws.cell(row=rec_row, column=1).font = bold_blue
        for line in [
            "L'AdE considera la via sostitutiva 26% obbligatoria per dividendi non",
            "qualificati percepiti tramite intermediario non residente",
            "(art. 27 c. 4-bis DPR 600/1973). La via RL+CE (Cass. 35454/2022) e'",
            "difendibile ma zona grigia.",
            "Gli interessi (tipo 'A') seguono regole analoghe: sostitutiva 26%",
            "obbligatoria art. 26 DPR 600 -> RM31.",
        ]:
            ws.append([line])
            ws.cell(row=ws.max_row, column=1).font = italic_dim

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 60


def _write_rw(ws: Worksheet, report: TaxReport) -> None:
    headers = [
        "Cod.",
        "ISIN",
        "Simbolo",
        "Azienda",
        "Descrizione",
        "Valuta",
        "Paese",
        "Quantita",
        "Acquisto",
        "Vendita",
        "Val. iniz. orig.",
        "Val. fin. orig.",
        "Cambio iniz.",
        "Cambio fin.",
        "Val. iniz. EUR",
        "Val. fin. EUR",
        "Giorni",
        "Quota %",
        "IVAFE",
    ]
    _write_header(ws, headers)

    for line in report.rw_lines:
        row = [
            line.codice_investimento,
            line.isin,
            line.symbol,
            line.long_description,
            line.description,
            line.currency,
            line.country,
            float(line.quantity),
            line.acquisition_date.isoformat() if line.acquisition_date else "",
            line.disposed_date.isoformat() if line.disposed_date else "",
            float(line.initial_value),
            float(line.final_value),
            float(line.ecb_rate_initial),
            float(line.ecb_rate_final),
            float(line.initial_value_eur),
            float(line.final_value_eur),
            line.days_held,
            float(line.ownership_pct),
            float(line.ivafe_due),
        ]
        ws.append(row)

    ws.append([])
    total_row = ws.max_row + 1
    ws.append(
        [
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "TOTALE",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            float(report.total_ivafe),
        ]
    )
    ws.cell(row=total_row, column=19).number_format = _MONEY_FMT
    ws.cell(row=total_row, column=19).font = _HEADER_FONT

    _format_money_columns(ws, [11, 12, 15, 16, 19], 2, ws.max_row)
    _auto_width(ws)


def _write_rt(ws: Worksheet, report: TaxReport) -> None:
    headers = [
        "Simbolo",
        "ISIN",
        "Azienda",
        "Data acquisto",
        "Data vendita",
        "Quantita",
        "Valuta",
        "Corrispettivo EUR",
        "Commissione (val.)",
        "Commissione EUR",
        "Costo VN EUR",
        "Costo VN (val.)",
        "Costo VN per azione",
        "Costo broker (val.)",
        "Costo broker per azione",
        "P/L EUR",
        "Cambio BCE",
        "Forex",
        "P/L broker (val.)",
        "P/L broker EUR",
    ]
    _write_header(ws, headers)

    for line in report.rt_lines:
        qty = line.quantity
        vn_per_share = (
            float(line.normal_value_cost / qty) if qty and line.normal_value_cost else None
        )
        broker_per_share = (
            float(line.broker_cost_basis / qty) if qty and line.broker_cost_basis else None
        )
        ws.append(
            [
                line.symbol,
                line.isin,
                _strip_acquired(line.long_description),
                line.acquisition_date.isoformat(),
                line.sell_date.isoformat(),
                float(qty),
                line.currency,
                float(line.proceeds_eur),
                float(line.commission_native) if line.commission_native else None,
                float(line.commission_eur) if line.commission_eur else None,
                float(line.cost_basis_eur),
                float(line.normal_value_cost) if line.normal_value_cost else None,
                vn_per_share,
                float(line.broker_cost_basis) if line.broker_cost_basis else None,
                broker_per_share,
                float(line.gain_loss_eur),
                float(line.ecb_rate),
                "Si" if line.is_forex else "No",
                float(line.broker_pnl) if line.broker_pnl else None,
                float(line.broker_pnl_eur) if line.broker_pnl_eur else None,
            ]
        )

    ws.append([])
    total_row = ws.max_row + 1
    total_cost_eur = sum((line.cost_basis_eur for line in report.rt_lines), Decimal(0))
    total_proceeds_eur = sum((line.proceeds_eur for line in report.rt_lines), Decimal(0))
    total_comm_eur = sum((line.commission_eur for line in report.rt_lines), Decimal(0))
    broker_pnl_eur_total = sum(
        (line.broker_pnl_eur for line in report.rt_lines if not line.is_forex),
        Decimal(0),
    )
    netto_row: list[object] = [
        "",
        "",
        "",
        "",
        "",
        "",
        "NETTO",
        float(total_proceeds_eur),
        "",
        float(total_comm_eur) if total_comm_eur else "",
        float(total_cost_eur),
        "",
        "",
        "",
        "",
        float(report.net_capital_gain_loss),
        "",
        "",
        "",
        float(broker_pnl_eur_total),
    ]
    ws.append(netto_row)
    for col in (8, 10, 11, 16, 20):
        ws.cell(row=total_row, column=col).number_format = _MONEY_FMT
        ws.cell(row=total_row, column=col).font = _HEADER_FONT

    _format_money_columns(ws, [8, 9, 10, 11, 12, 13, 14, 15, 16, 19, 20], 2, ws.max_row)
    _auto_width(ws)


def _write_rl(ws: Worksheet, report: TaxReport) -> None:
    headers = [
        "Descrizione",
        "Valuta",
        "Lordo",
        "Lordo EUR",
        "Ritenuta",
        "Ritenuta EUR",
        "Netto EUR",
    ]
    _write_header(ws, headers)

    for line in report.rl_lines:
        ws.append(
            [
                line.description,
                line.currency,
                float(line.gross_amount),
                float(line.gross_amount_eur),
                float(line.wht_amount),
                float(line.wht_amount_eur),
                float(line.net_amount_eur),
            ]
        )

    ws.append([])
    total_row = ws.max_row + 1
    ws.append(
        [
            "",
            "",
            "TOTALI",
            float(report.total_gross_interest_eur),
            "",
            float(report.total_wht_eur),
            float(report.total_gross_interest_eur - report.total_wht_eur),
        ]
    )
    for col in (4, 6, 7):
        ws.cell(row=total_row, column=col).number_format = _MONEY_FMT
        ws.cell(row=total_row, column=col).font = _HEADER_FONT

    _format_money_columns(ws, [3, 4, 5, 6, 7], 2, ws.max_row)
    _auto_width(ws)


def _write_forex(ws: Worksheet, report: TaxReport) -> None:
    ws.append(["Analisi Soglia Valutaria", "", f"Anno fiscale {report.tax_year}"])
    ws["A1"].font = Font(bold=True, size=12)
    ws.append(
        [
            "Soglia: EUR 51.645,69",
            "",
            f"Risultato: {'SUPERATA' if report.forex_threshold_breached else 'NON SUPERATA'}",
            "",
            f"Massimo consecutivo: {report.forex_max_consecutive_days} giorni",
        ]
    )
    ws.append([])

    headers = [
        "Data",
        "Saldo USD",
        "Equiv. EUR",
        "Cambio",
        "Giorno lavorativo",
        "Sopra soglia",
    ]
    _write_header(ws, headers, start_row=4)

    for rec in report.forex_daily_records:
        if rec.usd_balance == 0 and not rec.above_threshold:
            continue  # skip zero-balance days to keep sheet manageable
        ws.append(
            [
                rec.date.isoformat(),
                float(rec.usd_balance),
                float(rec.eur_equivalent),
                float(rec.fx_rate),
                "Si" if rec.is_business_day else "",
                "SI" if rec.above_threshold else "",
            ]
        )

    _format_money_columns(ws, [2, 3], 5, ws.max_row)
    _auto_width(ws)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _write_header(ws: Worksheet, headers: list[str], start_row: int = 1) -> None:
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")


def _format_money_columns(ws: Worksheet, columns: list[int], start_row: int, end_row: int) -> None:
    for col in columns:
        for row in range(start_row, end_row + 1):
            cell = ws.cell(row=row, column=col)
            if isinstance(cell.value, (int, float)):
                cell.number_format = _MONEY_FMT


def _auto_width(ws: Worksheet) -> None:
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for row in range(1, min(ws.max_row + 1, 50)):  # sample first 50 rows
            cell = ws.cell(row=row, column=col_idx)
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 40)
